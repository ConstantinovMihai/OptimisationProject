#=============================================================================================================================
# Created on 11 nov. 2019
# 
# @author: Bruno F. Santos
#
# Code for the Multi-commodity Flows problem (Arc version) for lectures of course AE4423 (Airline Planning & Optimisation)
#=============================================================================================================================



from gurobipy import *
from numpy import *
from openpyxl import *
from time import *


class Arc:
    def __init__(self,origin, destination, cost, capacity):
        self.From   = origin
        self.To     = destination
        self.Cost   = cost
        self.Capac  = capacity

class Commodity:
    def __init__(self,origin, destination, quantity):
        self.From   = origin
        self.To     = destination
        self.Quant  = quantity

class Node:
    def __init__(self):
        self.InLinks  = [ ]         #List of Nodes connected to self via a direct link
        self.OutLinks = [ ]         #List of Nodes connected from self via a direct link
    
    def addInLink(self,Node):
        self.InLinks.append(Node)       
    
    def addOutLink(self,Node):
        self.OutLinks.append(Node)



def MCF_Problem (Arcs, Nodes, Commodities):
    "Solve the MCF Problem with a Arc-based formulation using linear programming."
    
    model = Model("MCF")                # LP model (this is an object)
    
    x = {}                              # Decision Variables (DVs)
    for m in range(len(Arcs)):
        for k in range(len(Commodities)):
            x[k,Arcs[m].From,Arcs[m].To] = model.addVar(obj=Arcs[m].Cost, vtype ="C",
                                                  name = "x(%d,%d,%d)"%(k,Arcs[m].From,Arcs[m].To))
    
    model.update()                      # update model with the DVs before adding constraints

    Continuity = {}                       # build 'continuity' constraints
    for k in range(len(Commodities)):
        for j in range(len(Nodes)):
            if j == Commodities[k].From:
                Continuity[k,j] = model.addConstr(quicksum(x[k,j,p] for p in Nodes[j].OutLinks) - quicksum(x[k,p,j] for p in Nodes[j].InLinks),
                                        '=', Commodities[k].Quant, name ='Continuity(%d,%d)' %(k,j) )
            elif j == Commodities[k].To:
                Continuity[k,j] = model.addConstr(quicksum(x[k,j,p] for p in Nodes[j].OutLinks) - quicksum(x[k,p,j] for p in Nodes[j].InLinks),
                                        '=', -Commodities[k].Quant, name ='Continuity(%d,%d)' %(k,j) )
            else:
                Continuity[k,j] = model.addConstr(quicksum(x[k,j,p] for p in Nodes[j].OutLinks) - quicksum(x[k,p,j] for p in Nodes[j].InLinks),
                                        '=', 0, name ='Continuity(%d,%d)' %(k,j) )


    Capacity = {}                       # build 'capacity' constraints
    for m in range(len(Arcs)):
        Capacity [Arcs[m].From,Arcs[m].To] = model.addConstr(quicksum(x[k,Arcs[m].From,Arcs[m].To] for k in range(len(Commodities))),
                                                             '<=', Arcs[m].Capac, name = 'Capacity(%d)' %(m))

    model.update()
    model.write("MCF_Model.lp")
    model.optimize()
    
    status = model.status
    if status != GRB.Status.OPTIMAL:
        if status == GRB.Status.UNBOUNDED:
            print('The model cannot be solved because it is unbounded')
        elif status == GRB.Status.INFEASIBLE:
            print('The model is infeasible; computing IIS')
            model.computeIIS()
            print('\nThe following constraint(s) cannot be satisfied:')
            for c in model.getConstrs():
                if c.IISConstr:
                    print('%s' % c.constrName)
        elif status != GRB.Status.INF_OR_UNBD:
            print('Optimization was stopped with status %d' % status)
        exit(0)


    print
    for k in range(len(Arcs)):
        Flow = 0
        for m in range(len(Commodities)):
            Flow += x[m,Arcs[k].From,Arcs[k].To].X
        if int(Flow)>0:
            print ("Arc(%d,%d) \t" %(Arcs[k].From + 1,Arcs[k].To + 1), int(Flow))
            print
    print
    print ("Objective Function =", model.ObjVal/1.0)
    print ("------------------------------------------------------------------------")
    
    
    

if __name__ == '__main__':
    #=================================================================================================
    # Input excel file with arcs data (sheet1) and commodities data (sheet2)
    Arcs        = []
    Nodes       = []
    Commodities = []
    
    wb = load_workbook("Input_Lectures.xlsx", read_only=True)
    List_arcs = tuple(wb["Arcs"].iter_rows())
    List_commo = tuple(wb["Commodities"].iter_rows())
        
    # Generate Nodes and info about SP
    NewNode = []
    NodesNames = []
    for (ID,origin,destination,Cost,Capacity) in List_arcs[1:]:
        NewNode.append(int(origin.value))
        NewNode.append(int(destination.value))
    NewNode=set(NewNode)    #eliminate repeated nodes in the list
    for i in range(len(NewNode)):
        new = Node()
        Nodes.append(new)
    


    # Insert arcs (and their characteristics) in a list of Arcs
    for (ID, origin, destination,Cost,Capacity) in List_arcs[1:]:
        new = Arc (int(origin.value), int(destination.value), int(Cost.value), int(Capacity.value))
        Arcs.append(new)
        Nodes[int(origin.value)].addOutLink(new.To)
        Nodes[int(destination.value)].addInLink(new.From)

    # Insert commodities (and their characteristics) in a list of Commodities
    for (ID, origin, destination,quantity) in List_commo[1:]:
        new = Commodity (int(origin.value), int(destination.value), int(quantity.value))
        Commodities.append(new)

    
    del new, NewNode
    #=================================================================================================
    start_time = time()
    # RUN MCF PROBLEM
    MCF_Problem(Arcs, Nodes, Commodities)
    
    elapsed_time = time() - start_time

    print ("Run Time = ", elapsed_time)
    print ("END")
